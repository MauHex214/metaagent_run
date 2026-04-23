"""CDE 构建子模块。

流程：
  Step A: 自动生成 Tier 1 CDE（查 MIxS xlsx）
  Step B: LLM 批量建议 Tier 2 CDE
  Step C: 合并产出 cde_per_environment.json

测通阶段不需要人工审核，build-cde 默认直接合并 LLM 建议（每条标
`reviewed_by_human: false` 警示）。
"""

import asyncio
import json
import logging
import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional

from metaagent_run.core import (
    AsyncLocalModelClient,
    backoff_with_jitter,
    detect_truncation,
    continue_json_until_ok,
    extract_json_from_response_with_repair,
)

from .config import RuntimeConfig

LOGGER = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════
#  MIxS xlsx 解析
# ═══════════════════════════════════════════════════════════

def _parse_units(cell: Any) -> List[str]:
    if not cell:
        return []
    return [u.strip() for u in str(cell).split(",") if u.strip()]


def load_mixs_syntax_table(xlsx_path: Path) -> Dict[str, Dict[str, Any]]:
    """读 mixs_v6.xlsx，返回 {slot_name: {value_syntax, preferred_unit, example, mixs_id}}.

    只使用 environmental_packages sheet 中 package ∈ {water, sediment} 的行。
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 先读 MIxS 主表：除了 mixs_id，还提取 value_syntax/preferred_unit/example
    # 主表覆盖核心通用字段（collection_date / geo_loc_name / lat_lon 等）
    id_lookup: Dict[str, str] = {}
    result: Dict[str, Dict[str, Any]] = {}
    if "MIxS" in wb.sheetnames:
        mixs_sheet = wb["MIxS"]
        headers = [c.value for c in mixs_sheet[1]]
        try:
            name_idx = headers.index("Structured comment name")
            id_idx = headers.index("MIXS ID")
            syntax_idx = headers.index("Value syntax")
            unit_idx = headers.index("Preferred unit")
            example_idx = headers.index("Example")
            for row in mixs_sheet.iter_rows(min_row=2, values_only=True):
                name = row[name_idx]
                if not name:
                    continue
                name_str = str(name)
                id_lookup[name_str] = str(row[id_idx]) if row[id_idx] else ""
                result[name_str] = {
                    "value_syntax": str(row[syntax_idx] or ""),
                    "preferred_unit": _parse_units(row[unit_idx]),
                    "example": str(row[example_idx] or ""),
                    "mixs_id": id_lookup[name_str],
                }
        except ValueError:
            LOGGER.warning("MIxS sheet missing expected columns")

    if "environmental_packages" not in wb.sheetnames:
        LOGGER.error("environmental_packages sheet not found")
        return result

    env_sheet = wb["environmental_packages"]
    env_headers = [c.value for c in env_sheet[1]]
    try:
        pkg_idx = env_headers.index("Environmental package")
        name_idx = env_headers.index("Structured comment name")
        syntax_idx = env_headers.index("Value syntax")
        unit_idx = env_headers.index("Preferred unit")
        example_idx = env_headers.index("Example")
    except ValueError as e:
        LOGGER.error("environmental_packages sheet missing column: %s", e)
        return {}

    # 环境包行覆盖/补充主表字段（环境包通常给出更具体的 preferred_unit）
    for row in env_sheet.iter_rows(min_row=2, values_only=True):
        if row[pkg_idx] not in ("water", "sediment"):
            continue
        name = row[name_idx]
        if not name:
            continue
        name_str = str(name)
        new_units = _parse_units(row[unit_idx])
        if name_str in result:
            # 同字段已存在（来自主表或另一环境包）→ 并集 preferred_unit
            existing = result[name_str].get("preferred_unit", [])
            merged = list(dict.fromkeys(existing + new_units))
            result[name_str]["preferred_unit"] = merged
            # value_syntax 以环境包的为准（更精确）
            if row[syntax_idx]:
                result[name_str]["value_syntax"] = str(row[syntax_idx])
            continue
        result[name_str] = {
            "value_syntax": str(row[syntax_idx] or ""),
            "preferred_unit": new_units,
            "example": str(row[example_idx] or ""),
            "mixs_id": id_lookup.get(name_str, ""),
        }
    LOGGER.info(
        "Parsed MIxS (main + water + sediment): %d unique fields", len(result),
    )
    return result


# ═══════════════════════════════════════════════════════════
#  Normalizer 选择规则
# ═══════════════════════════════════════════════════════════

def pick_normalizer(value_syntax: str) -> str:
    """根据 value_syntax 字符串返回 normalizer 标识。"""
    s = (value_syntax or "").strip()
    exact = {
        "{float} {unit}": "float_with_unit",
        "{float}": "float_only",
        "{integer}": "integer",
        "{boolean}": "boolean",
        "{timestamp}": "timestamp",
        "{float} {float}": "lat_lon",
        "{termLabel} {[termID]}": "ontology_term",
        "{duration}": "duration",
        "{text}": "passthrough",
        "{float}:{float}": "ratio",
    }
    if s in exact:
        return exact[s]
    if s.startswith("[") and "|" in s and s.endswith("]"):
        return "enum"
    if s.startswith("{text};{float} {unit}"):
        return "compound_text_measurement"
    if "-" in s and "float" in s and "unit" in s:
        return "range"
    return "passthrough"


# ═══════════════════════════════════════════════════════════
#  Tier 1 自动生成
# ═══════════════════════════════════════════════════════════

def build_tier1_autogen(
    env_targets: Dict[str, Any],
    mixs_syntax: Dict[str, Dict[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    """对 per_environment 下每个 env 的 Tier 1 字段，查 MIxS 表生成 CDE 条目。"""
    result: Dict[str, Dict[str, Any]] = {}
    per_env = env_targets.get("per_environment", {})
    missing: List[str] = []
    for env, env_data in per_env.items():
        env_cde: Dict[str, Any] = {}
        for f in env_data.get("fields", []):
            if not isinstance(f, dict):
                continue
            if not (f.get("tier") == 1 and f.get("mapped")):
                continue
            slot = f.get("slot") or f.get("mixs_slot") or ""
            if not slot:
                continue
            mixs_info = mixs_syntax.get(slot)
            if not mixs_info:
                missing.append("%s/%s" % (env, slot))
                continue
            env_cde["mixs:" + slot] = {
                "tier": 1,
                "mixs_id": mixs_info["mixs_id"],
                "value_syntax": mixs_info["value_syntax"],
                "preferred_unit": mixs_info["preferred_unit"],
                "example": mixs_info["example"],
                "normalizer": pick_normalizer(mixs_info["value_syntax"]),
                "reviewed_by_human": False,  # tier1 来自 MIxS 标准，无需 review，但保留字段
            }
        result[env] = env_cde
    if missing:
        LOGGER.warning(
            "%d Tier 1 fields not found in MIxS water+sediment packages (first 5: %s)",
            len(missing), missing[:5],
        )
    LOGGER.info(
        "Tier 1 autogen: %d envs, total entries=%d",
        len(result), sum(len(v) for v in result.values()),
    )
    return result


# ═══════════════════════════════════════════════════════════
#  Tier 2 LLM 建议
# ═══════════════════════════════════════════════════════════

TIER2_LLM_PROMPT = """You are a hydrosphere environmental metadata expert.

For each field below (extracted from scientific literature but NOT mapped to any MIxS standard),
suggest a normalization rule. Each field includes its name, optional aliases (synonyms),
and up to 3 sample evidence quotes from literature.

Reference MIxS Value syntax patterns:
  - "<float> <unit>"  — measurement with unit (most common; e.g., depth, temp)
  - "<float>"         — single number (e.g., pH)
  - "<integer>"       — integer count
  - "<boolean>"       — true/false
  - "<timestamp>"     — ISO 8601 datetime
  - "<float> <float>" — lat/lon pair
  - "<termLabel> [<termID>]" — ENVO/OBO ontology term
  - "<float>-<float> <unit>" — range
  - "[opt1|opt2|...]" — enumeration
  - "<duration>"      — time duration
  - "<text>"          — free text
  - "<text>;<float> <unit>" — compound

Note: in your output, use the actual MIxS curly-brace forms (e.g., "{float} {unit}"),
not the angle-bracket placeholder shown above (the angle brackets here only avoid
formatting conflicts in this prompt).

Choose the most appropriate Value syntax for each field, suggest preferred units (use a list,
ordered by likelihood), and provide a realistic example.

Available normalizer names (use exactly):
  passthrough, float_with_unit, float_only, integer, boolean, timestamp, lat_lon,
  ontology_term, range, enum, duration, compound_text_measurement, ratio

Fields to process:
{FIELDS_SECTION}

OUTPUT FORMAT (STRICT JSON):
{
  "results": [
    {
      "field": "<original field name>",
      "value_syntax": "<MIxS-style pattern using curly braces>",
      "preferred_unit": ["unit1", "unit2"],
      "example": "<realistic value example>",
      "normalizer": "<one of the normalizer names>",
      "llm_reasoning": "<one-line explanation>"
    }
  ]
}

Output the JSON object only, no surrounding prose. End your response with </json>.
"""


def _build_fields_section(
    fields: List[Dict[str, Any]],
    schema_contexts: Dict[str, List[Dict[str, str]]],
    evidence_per_field: int,
) -> str:
    lines: List[str] = []
    for f in fields:
        name = f.get("field", "")
        aliases = f.get("aliases", [])
        contexts = schema_contexts.get(name, [])[:evidence_per_field]
        lines.append("- name: %s" % name)
        if aliases:
            lines.append("  aliases: %s" % ", ".join(aliases))
        if contexts:
            lines.append("  evidence:")
            for ctx in contexts:
                snippet = (ctx.get("text", "") or "")[:200].replace("\n", " ")
                lines.append("    - %s" % snippet)
        lines.append("")
    return "\n".join(lines)


async def _llm_suggest_batch(
    client: AsyncLocalModelClient,
    fields_batch: List[Dict[str, Any]],
    schema_contexts: Dict[str, Any],
    cfg: RuntimeConfig,
) -> List[Dict[str, Any]]:
    """对一批 Tier 2 字段调 LLM 产生建议。"""
    fields_section = _build_fields_section(
        fields_batch, schema_contexts, cfg.tier2_evidence_per_field,
    )
    # 用 replace 而非 format，避免 prompt 内 MIxS {float}/{unit} 等被误识别为占位符
    prompt = TIER2_LLM_PROMPT.replace("{FIELDS_SECTION}", fields_section)
    messages = [
        {"role": "system", "content": "You are an expert in environmental metadata standards (MIxS / GSC)."},
        {"role": "user", "content": prompt},
    ]
    last_error = None
    for attempt in range(cfg.retry_times):
        temp = cfg.retry_temps[min(attempt, len(cfg.retry_temps) - 1)]
        # 流式优先
        text: Optional[str] = None
        try:
            resp = await client.chat_streaming_with_signals(messages, temperature_override=temp)
            if resp:
                text = resp.get("text") if isinstance(resp.get("text"), str) else None
                if text:
                    status = detect_truncation(
                        text, bool(resp.get("saw_done", False)),
                        resp.get("finish_reason"), stop_sentinel=cfg.llm_stop_sentinel,
                    )
                    if status != "ok":
                        cont = await continue_json_until_ok(
                            client, messages, text, client.max_tokens,
                            max_tokens_cap=cfg.llm_max_tokens_cap,
                            stop_sentinel=cfg.llm_stop_sentinel,
                            max_rounds=cfg.continuation_max_rounds,
                        )
                        if cont:
                            text = cont
        except Exception as e:
            last_error = "stream exception: %s" % e
            text = None
        if not text:
            try:
                text = await client.chat(messages, temperature_override=temp)
            except Exception as e:
                last_error = "chat exception: %s" % e
                text = None
        if not text:
            await asyncio.sleep(backoff_with_jitter(attempt, base=cfg.backoff_base, cap=cfg.backoff_cap))
            continue
        parsed = extract_json_from_response_with_repair(
            text, stop_sentinel=cfg.llm_stop_sentinel, target_keys=("results",),
            enable_p0=True, enable_p1=False,
        )
        results = None
        if isinstance(parsed, dict):
            results = parsed.get("results")
        elif isinstance(parsed, list):
            for e in parsed:
                if isinstance(e, dict) and "results" in e:
                    results = e["results"]
                    break
        if isinstance(results, list):
            return results
        last_error = "no results array"
        await asyncio.sleep(backoff_with_jitter(attempt, base=cfg.backoff_base, cap=cfg.backoff_cap))
    LOGGER.warning("LLM tier2 suggest failed for batch (size=%d). Last: %s",
                   len(fields_batch), last_error)
    return []


def _normalize_tier2_suggestion(raw: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """清洗 LLM 输出的单条建议；不合格返回 None。"""
    if not isinstance(raw, dict):
        return None
    name = str(raw.get("field", "")).strip()
    if not name:
        return None
    syntax = str(raw.get("value_syntax", "{text}")).strip() or "{text}"
    units = raw.get("preferred_unit", [])
    if isinstance(units, str):
        units = [u.strip() for u in units.split(",") if u.strip()]
    elif isinstance(units, list):
        units = [str(u).strip() for u in units if str(u).strip()]
    else:
        units = []
    example = str(raw.get("example", "")).strip()
    normalizer = str(raw.get("normalizer", "")).strip() or pick_normalizer(syntax)
    reasoning = str(raw.get("llm_reasoning", "")).strip()
    return {
        "tier": 2,
        "value_syntax": syntax,
        "preferred_unit": units,
        "example": example,
        "normalizer": normalizer,
        "llm_reasoning": reasoning,
        "reviewed_by_human": False,
    }


async def build_tier2_suggestions(
    tier2_field_pool: Dict[str, List[Dict[str, Any]]],
    schema_contexts: Dict[str, Any],
    cfg: RuntimeConfig,
) -> Dict[str, Dict[str, Any]]:
    """对每个环境的 Tier 2 字段调 LLM 产出建议；按字段去重（同字段在多 env 出现取一份）.

    返回结构: {field_name: cde_entry}（不带 env 维度，因为 Tier 2 字段的归一化规则与环境无关）。
    """
    # 收集所有 Tier 2 字段（按 field name 去重）
    seen: Dict[str, Dict[str, Any]] = {}
    for env, fields in tier2_field_pool.items():
        for f in fields:
            name = f.get("field", "")
            if name and name not in seen:
                seen[name] = {"field": name, "aliases": []}
    if not seen:
        return {}
    fields_list = list(seen.values())
    LOGGER.info("Tier 2 unique field count: %d, batch size %d",
                len(fields_list), cfg.tier2_llm_batch_size)

    # 分批
    batches = [
        fields_list[i:i + cfg.tier2_llm_batch_size]
        for i in range(0, len(fields_list), cfg.tier2_llm_batch_size)
    ]
    suggestions: Dict[str, Dict[str, Any]] = {}
    async with AsyncLocalModelClient(
        base_url=cfg.llm_base_url, model=cfg.llm_model,
        temperature=cfg.llm_temperature, max_tokens=cfg.llm_max_tokens,
        api_key=cfg.llm_api_key, stop_sentinel=cfg.llm_stop_sentinel,
        api_style=cfg.llm_api_style, auth_mode=cfg.llm_auth_mode,
    ) as client:
        for i, batch in enumerate(batches):
            LOGGER.info("Tier 2 LLM batch %d/%d (size=%d)...",
                        i + 1, len(batches), len(batch))
            batch_results = await _llm_suggest_batch(client, batch, schema_contexts, cfg)
            for raw in batch_results:
                entry = _normalize_tier2_suggestion(raw)
                if entry:
                    name = entry.pop("_name", None) or raw.get("field", "")
                    if name:
                        suggestions[name] = entry
    LOGGER.info("Tier 2 suggestions returned: %d / %d", len(suggestions), len(seen))
    # 兜底：未返回的字段强制给 passthrough
    for name in seen:
        if name not in suggestions:
            suggestions[name] = {
                "tier": 2,
                "value_syntax": "{text}",
                "preferred_unit": [],
                "example": "",
                "normalizer": "passthrough",
                "llm_reasoning": "LLM did not return; default to passthrough",
                "reviewed_by_human": False,
            }
    return suggestions


# ═══════════════════════════════════════════════════════════
#  合并 + 写文件
# ═══════════════════════════════════════════════════════════

def merge_cde(
    tier1_per_env: Dict[str, Dict[str, Any]],
    tier2_suggestions: Dict[str, Dict[str, Any]],
    env_targets: Dict[str, Any],
) -> Dict[str, Dict[str, Any]]:
    """把 Tier 1 (按 env) + Tier 2 (按 field) 合并为 cde_per_environment.json 结构。"""
    result: Dict[str, Dict[str, Any]] = {}
    per_env = env_targets.get("per_environment", {})
    for env in per_env:
        env_cde = dict(tier1_per_env.get(env, {}))
        # 把该 env 的 Tier 2 字段加进来，key 用 internal:{slot}
        for f in per_env[env].get("fields", []):
            if not isinstance(f, dict):
                continue
            if f.get("tier") != 2:
                continue
            slot = f.get("slot") or f.get("field") or ""
            if not slot:
                continue
            entry = tier2_suggestions.get(slot)
            if not entry:
                continue
            env_cde["internal:" + slot] = entry
        result[env] = env_cde
    return result


# ═══════════════════════════════════════════════════════════
#  顶层入口
# ═══════════════════════════════════════════════════════════

def _load_env_targets(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError("env_targets not found: %s" % path)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _load_schema_contexts(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _collect_tier2_pool(env_targets: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    pool: Dict[str, List[Dict[str, Any]]] = {}
    for env, env_data in env_targets.get("per_environment", {}).items():
        pool[env] = [f for f in env_data.get("fields", [])
                     if isinstance(f, dict) and f.get("tier") == 2]
    return pool


async def build_cde_async(cfg: RuntimeConfig, auto_merge: bool = True) -> None:
    """主流程：autogen Tier 1 + LLM Tier 2 + 合并写文件。"""
    cfg.output_dir.mkdir(parents=True, exist_ok=True)

    env_targets = _load_env_targets(cfg.input_dir / cfg.env_targets_file)
    schema_contexts = _load_schema_contexts(cfg.input_dir / cfg.schema_contexts_file)
    mixs_syntax = load_mixs_syntax_table(cfg.input_dir / cfg.mixs_xlsx_file)

    # Step A
    tier1 = build_tier1_autogen(env_targets, mixs_syntax)
    tier1_path = cfg.output_dir / cfg.cde_tier1_autogen_file
    tier1_path.write_text(
        json.dumps(tier1, ensure_ascii=False, indent=2), encoding="utf-8",
    )
    LOGGER.info("Wrote Tier 1 autogen: %s", tier1_path)

    # Step B
    tier2_pool = _collect_tier2_pool(env_targets)
    tier2_suggestions = await build_tier2_suggestions(tier2_pool, schema_contexts, cfg)
    sugg_path = cfg.output_dir / cfg.cde_tier2_suggestions_file
    sugg_path.write_text(
        json.dumps(tier2_suggestions, ensure_ascii=False, indent=2), encoding="utf-8",
    )
    LOGGER.info("Wrote Tier 2 suggestions: %s", sugg_path)

    # Step C
    if auto_merge:
        merged = merge_cde(tier1, tier2_suggestions, env_targets)
        cde_path = cfg.output_dir / cfg.cde_file
        cde_path.write_text(
            json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8",
        )
        LOGGER.info("Wrote merged CDE (NO human review): %s", cde_path)
        LOGGER.warning(
            "Tier 2 entries are LLM-generated; mark `reviewed_by_human: false`. "
            "Run `--merge-only` after manual review for production use.",
        )


def build_cde(cfg: RuntimeConfig, auto_merge: bool = True) -> None:
    asyncio.run(build_cde_async(cfg, auto_merge=auto_merge))


def merge_only(cfg: RuntimeConfig, tier1_path: Path, tier2_path: Path) -> None:
    """只合并已有的 Tier 1 autogen + Tier 2 reviewed → cde_per_environment.json。"""
    if not tier1_path.exists():
        raise FileNotFoundError("Tier 1 file not found: %s" % tier1_path)
    if not tier2_path.exists():
        raise FileNotFoundError("Tier 2 file not found: %s" % tier2_path)
    with open(tier1_path, "r", encoding="utf-8") as f:
        tier1 = json.load(f)
    with open(tier2_path, "r", encoding="utf-8") as f:
        tier2 = json.load(f)
    env_targets = _load_env_targets(cfg.input_dir / cfg.env_targets_file)
    merged = merge_cde(tier1, tier2, env_targets)
    cde_path = cfg.output_dir / cfg.cde_file
    cde_path.write_text(
        json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8",
    )
    LOGGER.info("Wrote merged CDE: %s", cde_path)
